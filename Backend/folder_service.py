import os
import logging
import json
import datetime
import shutil
import tempfile
import ezdxf
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment
from werkzeug.utils import secure_filename
from werkzeug.datastructures import FileStorage
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import io
import math

# Configuration du logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Création d'une application Flask autonome
app = Flask(__name__)
CORS(app, resources={
    r"/*": {
        "origins": ["http://localhost:3000"],
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "supports_credentials": True,
        "expose_headers": ["Content-Disposition"],
        "max_age": 3600
    }
})

@app.route('/create-folder', methods=['POST'])
def create_folder():
    """Crée un dossier utilisateur basé sur l'email fourni dans la requête"""
    logger.info("Requête POST reçue pour créer un dossier utilisateur")
    logger.info(f"Headers: {request.headers}")
    
    try:
        # Récupérer les données de la requête
        data = request.get_json() or {}
        logger.info(f"Données reçues: {data}")
        
        email = data.get('email')
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({'error': 'Email non fourni'}), 400
        
        logger.info(f"Email de l'utilisateur: {email}")
        
        # Créer le chemin du dossier Ressources
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        # Vérifier si le dossier Ressources existe, sinon le créer
        if not os.path.exists(resource_dir):
            logger.info(f"Création du dossier Ressources: {resource_dir}")
            os.makedirs(resource_dir)
        
        # Créer le nom du dossier utilisateur basé sur l'email
        folder_name = email.split('@')[0]
        resource_path = os.path.join(resource_dir, folder_name)
        
        # Vérifier si le dossier existe déjà
        if os.path.exists(resource_path):
            logger.info(f"Le dossier existe déjà: {resource_path}")
            return jsonify({
                'message': 'Le dossier existe déjà',
                'folderName': folder_name,
                'folderExists': True
            }), 200
        
        # Créer le dossier
        os.makedirs(resource_path)
        logger.info(f"Dossier créé avec succès: {resource_path}")
        
        # Ajouter l'entrée dans la base de données en communiquant avec le service principal
        try:
            # Récupérer l'ID utilisateur à partir de l'email
            user_data = {
                'email': email,
                'folder_name': folder_name
            }
            
            # Appeler le service principal pour ajouter l'entrée dans la base de données
            response = requests.post(
                'http://localhost:5000/api/users/register-folder',
                json=user_data,
                headers={'Content-Type': 'application/json'}
            )
            
            if response.status_code == 200 or response.status_code == 201:
                logger.info(f"Entrée ajoutée dans la base de données pour le dossier: {folder_name}")
            else:
                logger.warning(f"Impossible d'ajouter l'entrée dans la base de données: {response.text}")
        except Exception as e:
            logger.error(f"Erreur lors de l'ajout de l'entrée dans la base de données: {str(e)}")
        
        return jsonify({
            'message': 'Dossier créé avec succès',
            'folderName': folder_name,
            'folderExists': True
        }), 201
    
    except Exception as e:
        logger.error(f"Erreur lors de la création du dossier: {str(e)}")
        return jsonify({'error': f'Erreur lors de la création du dossier: {str(e)}'}), 500

@app.route('/check-folder', methods=['POST'])
def check_folder():
    """Vérifie si un dossier utilisateur existe basé sur l'email fourni dans la requête"""
    logger.info("Requête POST reçue pour vérifier un dossier utilisateur")
    logger.info(f"Headers: {request.headers}")
    
    try:
        # Récupérer les données de la requête
        data = request.get_json() or {}
        logger.info(f"Données reçues: {data}")
        
        email = data.get('email')
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({'error': 'Email non fourni'}), 400
        
        logger.info(f"Email de l'utilisateur: {email}")
        
        # Créer le chemin du dossier Ressources
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        # Vérifier si le dossier Ressources existe
        if not os.path.exists(resource_dir):
            logger.info(f"Le dossier Ressources n'existe pas: {resource_dir}")
            return jsonify({
                'folderExists': False,
                'folderName': email.split('@')[0],
                'message': 'Le dossier Ressources n\'existe pas'
            }), 200
        
        # Vérifier si le dossier utilisateur existe
        folder_name = email.split('@')[0]
        resource_path = os.path.join(resource_dir, folder_name)
        folder_exists = os.path.exists(resource_path)
        
        logger.info(f"Vérification du dossier: {resource_path}, existe: {folder_exists}")
        return jsonify({
            'folderExists': folder_exists,
            'folderName': folder_name,
            'message': 'Dossier vérifié avec succès'
        }), 200
    
    except Exception as e:
        logger.error(f"Erreur lors de la vérification du dossier: {str(e)}")
        return jsonify({'error': f'Erreur lors de la vérification du dossier: {str(e)}'}), 500

def get_folder_structure(folder_path):
    """Récupère la structure des fichiers et dossiers à partir d'un chemin donné"""
    structure = {
        'folders': [],
        'files': []
    }
    
    try:
        if not os.path.exists(folder_path):
            logger.warning(f"Le dossier n'existe pas: {folder_path}")
            return structure
        
        for item in os.listdir(folder_path):
            item_path = os.path.join(folder_path, item)
            
            if os.path.isdir(item_path):
                # C'est un dossier
                sub_structure = get_folder_structure(item_path)  # Récupérer récursivement la structure du sous-dossier
                folder_info = {
                    'name': item,
                    'path': os.path.relpath(item_path, folder_path),
                    'last_modified': datetime.datetime.fromtimestamp(os.path.getmtime(item_path)).strftime('%Y-%m-%d %H:%M:%S'),
                    'sub_structure': sub_structure  # Ajouter la sous-structure au dossier
                }
                structure['folders'].append(folder_info)
            else:
                # C'est un fichier
                file_size = os.path.getsize(item_path)
                structure['files'].append({
                    'name': item,
                    'path': os.path.relpath(item_path, folder_path),
                    'size': file_size,
                    'size_formatted': format_file_size(file_size),
                    'last_modified': datetime.datetime.fromtimestamp(os.path.getmtime(item_path)).strftime('%Y-%m-%d %H:%M:%S')
                })
        
        # Trier les dossiers et fichiers par nom
        structure['folders'].sort(key=lambda x: x['name'].lower())
        structure['files'].sort(key=lambda x: x['name'].lower())
        
        return structure
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la structure du dossier: {str(e)}")
        return structure

def format_file_size(size_in_bytes):
    """Formate la taille d'un fichier en unités lisibles"""
    if size_in_bytes < 1024:
        return f"{size_in_bytes} B"
    elif size_in_bytes < 1024 * 1024:
        return f"{size_in_bytes / 1024:.2f} KB"
    elif size_in_bytes < 1024 * 1024 * 1024:
        return f"{size_in_bytes / (1024 * 1024):.2f} MB"
    else:
        return f"{size_in_bytes / (1024 * 1024 * 1024):.2f} GB"

@app.route('/get-folder-files', methods=['POST'])
def get_folder_files():
    """Récupère les fichiers du dossier utilisateur"""
    logger.info("Requête POST reçue pour récupérer les fichiers du dossier utilisateur")
    logger.info(f"Headers: {request.headers}")
    
    try:
        # Récupérer les données de la requête
        data = request.get_json() or {}
        logger.info(f"Données reçues: {data}")
        
        email = data.get('email')
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({'error': 'Email non fourni'}), 400
        
        logger.info(f"Email de l'utilisateur: {email}")
        
        # Créer le chemin du dossier Ressources
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        # Vérifier si le dossier Ressources existe
        if not os.path.exists(resource_dir):
            logger.info(f"Le dossier Ressources n'existe pas: {resource_dir}")
            return jsonify({
                'folders': [],
                'files': [],
                'message': 'Le dossier Ressources n\'existe pas'
            }), 200
        
        # Vérifier si le dossier utilisateur existe
        folder_name = email.split('@')[0]
        user_folder_path = os.path.join(resource_dir, folder_name)
        
        if not os.path.exists(user_folder_path):
            logger.info(f"Le dossier utilisateur n'existe pas: {user_folder_path}")
            return jsonify({
                'folders': [],
                'files': [],
                'message': 'Le dossier utilisateur n\'existe pas'
            }), 200
        
        # Récupérer la structure du dossier
        folder_structure = get_folder_structure(user_folder_path)
        logger.info(f"Structure du dossier récupérée: {json.dumps(folder_structure, indent=2)}")
        
        return jsonify(folder_structure), 200
    
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des fichiers: {str(e)}")
        return jsonify({'error': f'Erreur lors de la récupération des fichiers: {str(e)}'}), 500

@app.route('/transfer-files', methods=['POST'])
def transfer_files():
    """Transfère les fichiers dans le dossier utilisateur"""
    logger.info("Requête POST reçue pour le transfert de fichiers")
    logger.info(f"Headers: {request.headers}")
    
    try:
        # Vérifier si les fichiers sont présents dans la requête
        if 'file1' not in request.files or 'file2' not in request.files:
            logger.error("Deux fichiers sont requis")
            return jsonify({"error": "Deux fichiers sont requis pour le transfert"}), 400

        # Récupérer les fichiers et les noms de fichiers
        file1 = request.files['file1']
        file2 = request.files['file2']
        filename1 = request.form.get('filename1')
        filename2 = request.form.get('filename2')
        custom_folder_name = request.form.get('customFolderName')
        
        # Extraire l'email du formulaire ou de l'en-tête
        email = request.form.get('email')
        
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({"error": "Email non fourni"}), 400
        
        logger.info(f"Email de l'utilisateur: {email}")
        
        # Créer le chemin du dossier Ressources
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        # Vérifier si le dossier Ressources existe
        if not os.path.exists(resource_dir):
            os.makedirs(resource_dir)
            logger.info(f"Dossier Ressources créé: {resource_dir}")
        
        # Déterminer le nom du dossier utilisateur
        folder_name = email.split('@')[0]
        user_folder_path = os.path.join(resource_dir, folder_name)
        
        # Vérifier si le dossier utilisateur existe
        if not os.path.exists(user_folder_path):
            os.makedirs(user_folder_path)
            logger.info(f"Dossier utilisateur créé: {user_folder_path}")
        
        # Créer le dossier de transfert (sous-dossier dans le dossier utilisateur)
        if custom_folder_name:
            transfer_folder = os.path.join(user_folder_path, secure_filename(custom_folder_name))
        else:
            transfer_folder = os.path.join(user_folder_path, 'Transfert')
        
        # Vérifier si le dossier de transfert existe
        if not os.path.exists(transfer_folder):
            os.makedirs(transfer_folder)
            logger.info(f"Dossier de transfert créé: {transfer_folder}")
        
        # Sécuriser les noms de fichiers
        if not filename1:
            filename1 = secure_filename(file1.filename)
        else:
            filename1 = secure_filename(filename1)
        
        if not filename2:
            filename2 = secure_filename(file2.filename)
        else:
            filename2 = secure_filename(filename2)
        
        # Enregistrer les fichiers
        file1_path = os.path.join(transfer_folder, filename1)
        file2_path = os.path.join(transfer_folder, filename2)
        
        file1.save(file1_path)
        file2.save(file2_path)
        logger.debug(f"Fichiers sauvegardés: {file1_path}, {file2_path}")
        
        return jsonify({"message": f"Fichiers transférés avec succès dans {os.path.basename(transfer_folder)}"}), 200
    
    except Exception as e:
        logger.error(f"Erreur lors du transfert des fichiers: {str(e)}")
        return jsonify({"error": f"Erreur lors du transfert des fichiers: {str(e)}"}), 500

def extract_file_data(file):
    """Extrait les données d'un fichier DXF"""
    temp_file_path = None
    try:
        logger.info(f"Début de l'extraction pour le fichier : {file.filename}")
        
        # Créer un fichier temporaire avec un chemin explicite et mode binaire
        with tempfile.NamedTemporaryFile(delete=False, suffix='.dxf', mode='wb') as temp_file:
            file.save(temp_file)
            temp_file_path = temp_file.name
            logger.info(f"Fichier temporaire sauvegardé : {temp_file_path}")
            
        # Charger le fichier DXF à partir du chemin temporaire
        logger.info(f"Lecture du fichier DXF : {temp_file_path}")
        doc = ezdxf.readfile(temp_file_path)
        
        # Extraire les calques (layers)
        layers = [
            {
                "name": layer.dxf.name,
                "color": layer.dxf.color if layer.dxf.color != 0 else 'N/A',
                "lineweight": layer.dxf.lineweight if hasattr(layer.dxf, 'lineweight') else None
            }
            for layer in doc.layers 
            if not layer.dxf.name.startswith('*')  # Exclure les calques système
        ]
        
        # Extraire les entités (polylignes, lignes, cercles, arcs, texte, etc.)
        modelspace = doc.modelspace()
        
        polylines = []
        lines = []
        circles = []
        arcs = []
        texts = []
        
        for entity in modelspace.query('*'):
            dxftype = entity.dxftype()
            if dxftype == 'POLYLINE':
                polylines.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'vertices': [{'x': v[0], 'y': v[1]} for v in entity.points()],
                    'closed': entity.closed,
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
            elif dxftype == 'LWPOLYLINE':
                polylines.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'vertices': [{'x': v[0], 'y': v[1]} for v in entity.get_points()],
                    'closed': entity.closed,
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
            elif dxftype == 'LINE':
                lines.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'start': {'x': entity.dxf.start[0], 'y': entity.dxf.start[1]},
                    'end': {'x': entity.dxf.end[0], 'y': entity.dxf.end[1]},
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
            elif dxftype == 'CIRCLE':
                circles.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'center': {'x': entity.dxf.center[0], 'y': entity.dxf.center[1]},
                    'radius': entity.dxf.radius,
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
            elif dxftype == 'ARC':
                arcs.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'center': {'x': entity.dxf.center[0], 'y': entity.dxf.center[1]},
                    'radius': entity.dxf.radius,
                    'start_angle': entity.dxf.start_angle,
                    'end_angle': entity.dxf.end_angle,
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
            elif dxftype == 'TEXT':
                texts.append({
                    'type': dxftype,
                    'layer': entity.dxf.layer,
                    'text': entity.dxf.text,
                    'position': {'x': entity.dxf.insert[0], 'y': entity.dxf.insert[1]},
                    'height': entity.dxf.height,
                    'color': entity.dxf.color if entity.dxf.color != 0 else 'N/A',
                    'lineweight': entity.dxf.lineweight if hasattr(entity.dxf, 'lineweight') else None
                })
        
        # Statistiques globales
        total_entities = len(modelspace)
        statistics = {
            "layer_count": len(layers),
            "polyline_count": len(polylines),
            "line_count": len(lines),
            "circle_count": len(circles),
            "arc_count": len(arcs),
            "text_count": len(texts),
            "total_entities": total_entities
        }
        
        logger.info("Données extraites avec succès")
        return {
            "layers": layers,
            "polylines": polylines,
            "lines": lines,
            "circles": circles,
            "arcs": arcs,
            "texts": texts,
            "statistics": statistics
        }
    
    except Exception as e:
        logger.error(f"Erreur lors de l'extraction des données : {str(e)}")
        return {"error": f"Erreur lors de l'extraction des données : {str(e)}"}
    
    finally:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                logger.info(f"Fichier temporaire supprimé : {temp_file_path}")
            except Exception as e:
                logger.error(f"Erreur lors de la suppression du fichier temporaire : {str(e)}")

@app.route('/extract-data-from-file', methods=['POST'])
def extract_data_from_file():
    """Extrait les données d'un fichier DXF dans le dossier de l'utilisateur"""
    logger.info("Requête POST reçue pour extraire les données d'un fichier DXF")
    
    try:
        data = request.get_json() or {}
        logger.info(f"Données reçues: {data}")
        
        filename = data.get('filename')
        folder = data.get('folder', "")
        email = data.get('email')
        file_type = data.get('fileType', 'projet')
        
        logger.info(f"Type de fichier: {file_type}")
        
        if not filename:
            logger.error("Nom de fichier manquant")
            return jsonify({"error": "Nom de fichier requis"}), 400
        
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({"error": "Email non fourni"}), 400
        
        logger.info(f"Email de l'utilisateur: {email}")
        logger.info(f"Nom du fichier: {filename}")
        logger.info(f"Dossier: {folder}")
        
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        if not os.path.exists(resource_dir):
            logger.error(f"Le dossier Ressources n'existe pas: {resource_dir}")
            return jsonify({"error": "Dossier Ressources non trouvé"}), 400
        
        folder_name = email.split('@')[0]
        user_folder_path = os.path.join(resource_dir, folder_name)
        
        if not os.path.exists(user_folder_path):
            logger.error(f"Le dossier utilisateur n'existe pas: {user_folder_path}")
            return jsonify({"error": "Dossier utilisateur non trouvé"}), 400
        
        logger.info(f"Construction du chemin pour le fichier: {filename}, dossier: {folder}")
        
        if folder:
            file_path = os.path.join(user_folder_path, folder, filename)
        else:
            file_path = os.path.join(user_folder_path, filename)
            
        logger.info(f"Tentative d'accès au fichier: {file_path}")
        
        if os.path.exists(file_path):
            logger.info(f"Fichier trouvé: {file_path}")
        else:
            logger.warning(f"Fichier non trouvé au chemin exact: {file_path}, recherche alternative...")
            
            for root, dirs, files in os.walk(user_folder_path):
                if filename in files:
                    alt_file_path = os.path.join(root, filename)
                    logger.info(f"Fichier trouvé à un emplacement alternatif: {alt_file_path}")
                    file_path = alt_file_path
                    break
            
            if not os.path.exists(file_path):
                logger.error(f"Le fichier n'existe pas après recherche approfondie: {file_path}")
                found_files = []
                for root, dirs, files in os.walk(user_folder_path):
                    for file in files:
                        found_files.append(os.path.join(root, file))
                logger.info(f"Fichiers trouvés dans le dossier utilisateur: {found_files}")
                return jsonify({"error": "Fichier non trouvé"}), 400
        
        logger.info(f"Vérification du fichier: {file_path}")
        logger.info(f"Extension du fichier: {os.path.splitext(file_path)[1]}")
        
        if not file_path.lower().endswith('.dxf'):
            logger.warning(f"Format non standard détecté: {file_path} - tentative d'extraction quand même")
        
        with open(file_path, 'rb') as f:
            file_content = f.read()
        
        file_obj = FileStorage(
            stream=io.BytesIO(file_content),
            filename=os.path.basename(file_path),
            content_type='application/octet-stream'
        )
        
        extracted_data = extract_file_data(file_obj)
        logger.info(f"Données extraites avec succès pour {file_path} (Type: {file_type})")
        
        extracted_data['fileType'] = file_type
        
        rel_path = os.path.relpath(file_path, user_folder_path)
        logger.info(f"Chemin relatif du fichier: {rel_path}")
        
        path_components = rel_path.split(os.sep)
        
        if len(path_components) > 1:
            parent_folder = path_components[0]
            logger.info(f"Dossier parent détecté: {parent_folder}")
            extracted_data['sourcePath'] = rel_path
            extracted_data['parentFolder'] = parent_folder
        else:
            extracted_data['sourcePath'] = rel_path
            extracted_data['parentFolder'] = ''
        
        return jsonify(extracted_data), 200
    
    except Exception as e:
        logger.error(f"Erreur lors de l'extraction: {str(e)}")
        return jsonify({"error": f"Erreur lors de l'extraction: {str(e)}"}), 500

@app.route('/get-visa-content', methods=['POST'])
def get_visa_content():
    """Récupère le contenu d'un fichier visa.txt"""
    logger.info("Requête POST reçue pour récupérer le contenu d'un fichier visa")
    
    try:
        data = request.get_json() or {}
        file_path = data.get('filePath')
        email = data.get('email')
        
        logger.info(f"Tentative d'accès au fichier: {file_path}")
        
        if email and '/' in file_path and not file_path.startswith('/'):
            current_dir = os.path.dirname(os.path.abspath(__file__))
            resource_dir = os.path.join(current_dir, 'app', 'Ressources')
            folder_name = email.split('@')[0]
            user_folder_path = os.path.join(resource_dir, folder_name)
            
            complete_file_path = os.path.join(user_folder_path, file_path)
            logger.info(f"Chemin complet construit: {complete_file_path}")
            file_path = complete_file_path
        
        if not file_path or not os.path.exists(file_path):
            logger.error(f"Fichier non trouvé: {file_path}")
            
            if email:
                current_dir = os.path.dirname(os.path.abspath(__file__))
                resource_dir = os.path.join(current_dir, 'app', 'Ressources')
                folder_name = email.split('@')[0]
                user_folder_path = os.path.join(resource_dir, folder_name)
                
                if '/' in file_path:
                    filename = file_path.split('/')[-1]
                else:
                    filename = file_path
                    
                logger.info(f"Recherche du fichier {filename} dans le dossier utilisateur {user_folder_path}")
                
                found = False
                for root, dirs, files in os.walk(user_folder_path):
                    if filename in files:
                        file_path = os.path.join(root, filename)
                        found = True
                        logger.info(f"Fichier trouvé à: {file_path}")
                        break
                
                if not found:
                    logger.error(f"Fichier non trouvé après recherche approfondie")
                    return jsonify({'error': 'Fichier non trouvé'}), 404
            else:
                return jsonify({'error': 'Fichier non trouvé'}), 404
        
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        logger.info(f"Contenu récupéré avec succès: {file_path}")
        
        return jsonify({
            'message': 'Contenu récupéré avec succès',
            'content': content
        }), 200
    
    except Exception as e:
        logger.error(f"Erreur lors de la récupération du contenu: {str(e)}")
        return jsonify({'error': f'Erreur lors de la récupération du contenu: {str(e)}'}), 500

@app.route('/download-visa-file', methods=['GET'])
def download_visa_file():
    """Permet le téléchargement d'un fichier visa.txt"""
    logger.info("Requête GET reçue pour télécharger un fichier visa")
    
    try:
        file_path = request.args.get('filePath')
        email = request.args.get('email')
        
        logger.info(f"Paramètres reçus - filePath: {file_path}, email: {email}")
        
        if email and file_path:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            resource_dir = os.path.join(current_dir, 'app', 'Ressources')
            folder_name = email.split('@')[0]
            user_folder_path = os.path.join(resource_dir, folder_name)
            
            complete_file_path = os.path.join(user_folder_path, file_path)
            logger.info(f"Chemin complet construit: {complete_file_path}")
            file_path = complete_file_path
        
        if not file_path or not os.path.exists(file_path):
            logger.error(f"Fichier non trouvé: {file_path}")
            
            if email:
                current_dir = os.path.dirname(os.path.abspath(__file__))
                resource_dir = os.path.join(current_dir, 'app', 'Ressources')
                folder_name = email.split('@')[0]
                user_folder_path = os.path.join(resource_dir, folder_name)
                
                if '/' in file_path:
                    filename = file_path.split('/')[-1]
                else:
                    filename = file_path
                    
                logger.info(f"Recherche du fichier {filename} dans le dossier utilisateur {user_folder_path}")
                
                found = False
                for root, dirs, files in os.walk(user_folder_path):
                    if filename in files:
                        file_path = os.path.join(root, filename)
                        found = True
                        logger.info(f"Fichier trouvé à: {file_path}")
                        break
                
                if not found:
                    return jsonify({'error': 'Fichier non trouvé'}), 404
            else:
                return jsonify({'error': 'Fichier non trouvé'}), 404
        
        file_name = os.path.basename(file_path)
        
        logger.info(f"Téléchargement du fichier: {file_path}")
        return send_file(
            file_path,
            mimetype='text/plain',
            as_attachment=True,
            download_name=file_name
        )
    
    except Exception as e:
        logger.error(f"Erreur lors du téléchargement du fichier: {str(e)}")
        return jsonify({'error': f'Erreur lors du téléchargement du fichier: {str(e)}'}), 500

@app.route('/generate-visa-file', methods=['POST'])
def generate_visa_file():
    """Génère un fichier visa.txt avec les informations de surface calculées"""
    logger.info("Requête POST reçue pour générer un fichier visa.txt")
    
    try:
        data = request.get_json() or {}
        logger.info(f"Données reçues: {data}")
        
        email = data.get('email')
        surfaces = data.get('surfaces')
        floor_name = data.get('floorName', 'Sans nom')
        folder_path = data.get('folderPath', '')
        
        logger.info(f"Dossier parent pour le fichier visa: {folder_path}")
        
        if not email:
            logger.error("Email non fourni dans la requête")
            return jsonify({'error': 'Email non fourni'}), 400
        
        if not surfaces:
            logger.error("Données de surface non fournies")
            return jsonify({'error': 'Données de surface non fournies'}), 400
        
        current_dir = os.path.dirname(os.path.abspath(__file__))
        resource_dir = os.path.join(current_dir, 'app', 'Ressources')
        
        if not os.path.exists(resource_dir):
            logger.info(f"Création du dossier Ressources: {resource_dir}")
            os.makedirs(resource_dir)
        
        folder_name = email.split('@')[0]
        user_folder_path = os.path.join(resource_dir, folder_name)
        
        if not os.path.exists(user_folder_path):
            logger.info(f"Création du dossier utilisateur: {user_folder_path}")
            os.makedirs(user_folder_path)
        
        target_folder_path = user_folder_path
        
        if folder_path:
            target_folder_path = os.path.join(user_folder_path, folder_path)
            if not os.path.exists(target_folder_path):
                logger.warning(f"Le dossier {folder_path} n'existe pas, création: {target_folder_path}")
                os.makedirs(target_folder_path)
            else:
                logger.info(f"Utilisation du dossier existant: {target_folder_path}")
        
        output_folder_path = os.path.join(target_folder_path, 'Output')
        if not os.path.exists(output_folder_path):
            logger.info(f"Création du dossier Output: {output_folder_path}")
            os.makedirs(output_folder_path)
        
        file_name = f"visa_{floor_name.replace(' ', '_')}.txt"
        file_path = os.path.join(output_folder_path, file_name)
        
        content = generate_visa_content(surfaces, floor_name)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        logger.info(f"Fichier visa généré avec succès: {file_path}")
        
        return jsonify({
            'message': 'Fichier visa généré avec succès',
            'filePath': file_path
        }), 201
    
    except Exception as e:
        logger.error(f"Erreur lors de la génération du fichier visa: {str(e)}")
        return jsonify({'error': f'Erreur lors de la génération du fichier visa: {str(e)}'}), 500

@app.route('/generate-excel-file', methods=['POST'])
def generate_excel_file():
    """Génère un fichier Excel avec les informations de surface calculées"""
    logger.info("Requête POST reçue pour générer un fichier Excel")
    
    try:
        # Journal détaillé des étapes pour déboguer
        logger.info("Début du traitement de la requête generate-excel-file")
        
        # Détail complet de la requête
        try:
            data = request.get_json() or {}
            logger.info(f"Données reçues (brut): {data}")
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des données JSON: {str(e)}")
            return jsonify({'error': f'Erreur lors de la récupération des données JSON: {str(e)}'}), 500
        
        # Extraction et validation des données
        try:
            email = data.get('email', '')
            surfaces = data.get('surfaces', {})
            floor_name = data.get('floorName', 'Sans nom')
            folder_path = data.get('folderPath', '')
            
            logger.info(f"Données extraites - email: {email}, floor_name: {floor_name}, folder_path: {folder_path}")
            logger.info(f"Structure de 'surfaces': {list(surfaces.keys()) if isinstance(surfaces, dict) else 'NON-DICT'}")
            
            if not email:
                logger.error("Email non fourni dans la requête")
                return jsonify({'error': 'Email non fourni'}), 400
            
            if not surfaces:
                logger.error("Données des surfaces non fournies")
                return jsonify({'error': 'Données des surfaces non fournies'}), 400
        except Exception as e:
            logger.error(f"Erreur lors de la validation des données: {str(e)}")
            return jsonify({'error': f'Erreur lors de la validation des données: {str(e)}'}), 500
        
        # Préparation des dossiers
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            folder_name = email.split('@')[0]
            resource_dir = os.path.join(current_dir, 'app', 'Ressources', folder_name)
            
            logger.info(f"Chemins - current_dir: {current_dir}, folder_name: {folder_name}, resource_dir: {resource_dir}")
            
            if not os.path.exists(resource_dir):
                logger.warning(f"Le dossier de l'utilisateur n'existe pas: {resource_dir}, tentative de création")
                os.makedirs(resource_dir, exist_ok=True)
            
            if folder_path:
                project_dir = os.path.join(resource_dir, folder_path)
                if not os.path.exists(project_dir):
                    logger.info(f"Création du dossier projet: {project_dir}")
                    os.makedirs(project_dir, exist_ok=True)
                
                output_dir = os.path.join(project_dir, 'Output')
            else:
                output_dir = os.path.join(resource_dir, 'Output')
            
            if not os.path.exists(output_dir):
                logger.info(f"Création du dossier Output: {output_dir}")
                os.makedirs(output_dir, exist_ok=True)
                
            logger.info(f"Dossier Output finalisé: {output_dir} (existe: {os.path.exists(output_dir)})")
        except Exception as e:
            logger.error(f"Erreur lors de la préparation des dossiers: {str(e)}")
            return jsonify({'error': f'Erreur lors de la préparation des dossiers: {str(e)}'}), 500
        
        # Génération du fichier Excel
        try:
            # Sanitize du nom d'étage pour le nom de fichier
            sanitized_floor_name = ''.join(c if c.isalnum() or c in '-_' else '_' for c in floor_name)
            now = datetime.datetime.now()
            date_str = now.strftime('%Y%m%d_%H%M%S')
            excel_filename = f"surface_comparison_{sanitized_floor_name}_{date_str}.xlsx"
            excel_path = os.path.join(output_dir, excel_filename)
            
            logger.info(f"Préparation de la création du fichier Excel: {excel_path}")
            
            # Remplacer la fonction create_excel_document par une version très simplifiée pour tester
            # Cette version crée un fichier Excel minimal fonctionnel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "SDP"
            
            # En-têtes complets selon l'image de référence
            ws['A1'] = "Etages"
            ws['B1'] = "Destinations"
            ws['C1'] = "Surface existante avant travaux (A)"
            ws['D1'] = "Surface creee (B)"
            ws['E1'] = "Surface creee par changement de destination"
            ws['F1'] = "Surface demolie reconstruite"
            ws['G1'] = "Surface supprimee (D)"
            ws['H1'] = "Surface supprimee par changement de destination"
            ws['I1'] = "Surface projet"
            ws['J1'] = "Surface RDV"
            
            # Ajout de l'étage
            ws['A2'] = floor_name
            
            # Traitement très simplifié des destinations
            row = 2
            
            # Fonction pour calculer l'intersection entre deux polylignes
            def calculate_intersection(poly1, poly2):
                try:
                    # Convertir les sommets en objets Point
                    from shapely.geometry import Polygon
                    
                    # Créer les polygones à partir des sommets
                    poly1_pts = [(v['x'], v['y']) for v in poly1.get('vertices', [])]
                    poly2_pts = [(v['x'], v['y']) for v in poly2.get('vertices', [])]
                    
                    if len(poly1_pts) < 3 or len(poly2_pts) < 3:
                        return 0.0
                        
                    poly1_shapely = Polygon(poly1_pts)
                    poly2_shapely = Polygon(poly2_pts)
                    
                    # Vérifier si les polygones sont valides
                    if not poly1_shapely.is_valid or not poly2_shapely.is_valid:
                        return 0.0
                    
                    # Calculer l'intersection
                    intersection = poly1_shapely.intersection(poly2_shapely)
                    
                    # Retourner l'aire de l'intersection
                    return intersection.area
                    
                except Exception as e:
                    logger.warning(f"Erreur lors du calcul d'intersection: {str(e)}")
                    return 0.0
            
            # Dictionnaire de correspondance pour les cas spéciaux
            special_cases = {
                "AUTRE_BUREAU": "Autre bureau",
                "AUTRE_CONGRE": "Autre congrès exposition",
                "AUTRE_ENTREP": "Autre entrepôt",
                "AUTRE_INDUST": "Autre industrie",
                "COMMERCE_ART": "Commerce artisanat",
                "COMMERCE_AUT": "Commerce autre hébergement touristique",
                "COMMERCE_CIN": "Commerce cinéma",
                "COMMERCE_DE_": "Commerce de gros",
                "COMMERCE_HOT": "Commerce hôtel",
                "COMMERCE_RES": "Commerce restauration",
                "COMMERCE_SER": "Commerce service accueil clientèle",
                "EXPLOITATIO": "Exploitation forestière",
                "EXPLOITATION": "Exploitation agricole",
                "HABITATION_H": "Habitation hébergement",
                "HABITATION_L": "Habitation logement",
                "SPIC_ADMINIS": "Spic administration",
                "SPIC_ART_SPE": "Spic art spectacle",
                "SPIC_AUTRE": "Spic autre",
                "SPIC_ENSEIGN": "Spic enseignement santé",
                "SPIC_LT": "Spic lt",
                "SPIC_SPORT": "Spic sport"
            }
            
            # Journaliser la structure complète des données pour déboguer
            logger.info(f"Structure détaillée des surfaces: {json.dumps(surfaces, default=str)}")
            
            # Analyse détaillée de la structure des données
            existant_polylines = surfaces.get('existant', {}).get('polylines', [])
            projet_polylines = surfaces.get('projet', {}).get('polylines', [])
            
            logger.info(f"Existant polylines: {len(existant_polylines)}")
            logger.info(f"Projet polylines: {len(projet_polylines)}")
            
            # IMPORTANT: Rappel de l'inversion des fichiers
            # existant_polylines contient les données du fichier "Projet_demoli_feuille_TA.dxf" (surface existante avant travaux)
            # projet_polylines contient les données du fichier "Existant_exmple_demoli.dxf" (surface projet)
            
            # Fonction pour extraire le nom de destination d'une polyligne avec logging
            def get_destination_from_layer(layer):
                if not isinstance(layer, str):
                    logger.warning(f"Layer n'est pas une chaîne: {layer}")
                    return None
                    
                # Vérifier si c'est un calque SDP_1
                if 'GEX_EDS_SDP_1' not in layer:
                    return None
                
                # Journaliser le calque pour débogage
                logger.info(f"Traitement du calque: {layer}")
                
                # Extraire la partie après GEX_EDS_SDP_1-
                try:
                    parts = layer.split('GEX_EDS_SDP_1-')
                    if len(parts) >= 2:
                        raw_destination = parts[1]
                        
                        # Journaliser la destination brute extraite
                        logger.info(f"Destination brute extraite: {raw_destination}")
                        
                        # Nettoyer le nom de la destination de tout caractère indésirable à la fin
                        if raw_destination.endswith('_'):
                            raw_destination = raw_destination[:-1]
                        
                        # Cas spécial pour EXPLOITATIO0 qui apparaît dans votre fichier
                        if "EXPLOITATIO0" in raw_destination:
                            raw_destination = "EXPLOITATIO"
                        
                        # Tentative de correspondance avec les clés connues
                        for key in special_cases.keys():
                            if key == raw_destination:
                                logger.info(f"Match exact: {key}")
                                return key
                        
                        # Si pas de correspondance exacte, faire une recherche partielle
                        for key in special_cases.keys():
                            if key in raw_destination or raw_destination in key:
                                logger.info(f"Match partiel: {key} dans {raw_destination}")
                                return key
                        
                        logger.info(f"Destination finale: {raw_destination}")
                        return raw_destination
                except Exception as e:
                    logger.error(f"Erreur lors de l'extraction de la destination: {str(e)}")
                
                return None
            
            # Fonction pour déterminer si une polyligne est un calque spécial à déduire
            def is_special_layer(layer):
                if not isinstance(layer, str):
                    return False
                    
                patterns = ['GEX_EDS_SDP_2', 'GEX_EDS_SDP_3', 'GEX_EDS_SDP_4', 'GEX_EDS_SDP_5', 'GEX_EDS_SDP_7']
                return any(pattern in layer for pattern in patterns)
            
            # Fonction pour calculer la surface d'une polyligne
            def calculate_area(polyline):
                vertices = polyline.get('vertices', [])
                if len(vertices) < 3:
                    return 0
                    
                # Calculer l'aire en utilisant la formule de Shoelace
                area = 0.0
                n = len(vertices)
                for i in range(n):
                    j = (i + 1) % n
                    area += vertices[i]['x'] * vertices[j]['y']
                    area -= vertices[j]['x'] * vertices[i]['y']
                    
                return abs(area) / 2.0
            
            # Fonction pour déterminer si une polyligne est contenue dans une autre
            def is_contained(polyline1, polyline2):
                try:
                    vertices1 = polyline1.get('vertices', [])
                    vertices2 = polyline2.get('vertices', [])
                    
                    if not vertices1 or not vertices2:
                        return False
                        
                    # Trouver les limites de la boîte englobante pour polyline2
                    min_x2 = min(v['x'] for v in vertices2)
                    max_x2 = max(v['x'] for v in vertices2)
                    min_y2 = min(v['y'] for v in vertices2)
                    max_y2 = max(v['y'] for v in vertices2)
                    
                    # Vérifier si tous les points de polyline1 sont à l'intérieur de la boîte englobante de polyline2
                    points_inside = 0
                    for v in vertices1:
                        if min_x2 <= v['x'] <= max_x2 and min_y2 <= v['y'] <= max_y2:
                            points_inside += 1
                    
                    # Si au moins la moitié des points sont à l'intérieur, considérer comme contenu
                    return points_inside >= len(vertices1) / 2
                except Exception as e:
                    logger.warning(f"Erreur lors de la vérification de contenance: {str(e)}")
                    return False
                    
            # Structure pour stocker les résultats
            calculation_results = {
                'existant': {},       # Pour les surfaces existantes
                'projet': {},          # Pour les surfaces projet
                'cree_changement': {}, # Surface créée par changement de destination
                'demolie': {},         # Surface démolie reconstruite
                'supprimee': {},       # Surface supprimée (D)
                'supprimee_changement': {}, # Surface supprimée par changement de destination
                'demolition': {}       # Pour les surfaces de démolition par destination
            }
            
            # Identifier les polylignes de démolition (GEX_EDS_TA_SDP_CAHIER_DEMO)
            demolition_polylines = [p for p in existant_polylines 
                                 if 'GEX_EDS_TA_SDP_CAHIER_DEMO' in p.get('layer', '')]
            
            logger.info(f"Nombre de polylignes de démolition trouvées: {len(demolition_polylines)}")
            
            # Traiter les polylignes existantes (Projet_demoli_feuille_TA.dxf)
            main_existant_polylines = []
            special_existant_polylines = []
            
            for polyline in existant_polylines:
                layer = polyline.get('layer', '')
                if 'GEX_EDS_SDP_1' in layer:
                    destination = get_destination_from_layer(layer)
                    if destination:
                        main_existant_polylines.append(polyline)
                        area = calculate_area(polyline)
                        if destination not in calculation_results['existant']:
                            calculation_results['existant'][destination] = 0
                            calculation_results['demolition'][destination] = 0.0
                        calculation_results['existant'][destination] += area
                        logger.info(f"Existant: Destination {destination} - ajout surface {area}")
                        
                        # Calculer l'intersection avec les zones de démolition
                        for demo_poly in demolition_polylines:
                            intersection_area = calculate_intersection(polyline, demo_poly)
                            if intersection_area > 0:
                                calculation_results['demolition'][destination] += intersection_area
                                logger.info(f"  - Intersection avec zone de démolition: {intersection_area:.2f} m²")
                                
                elif is_special_layer(layer):
                    special_existant_polylines.append(polyline)
            
            # Traiter les polylignes projet (Existant_exmple_demoli.dxf)
            main_projet_polylines = []
            special_projet_polylines = []
            
            for polyline in projet_polylines:
                layer = polyline.get('layer', '')
                if 'GEX_EDS_SDP_1' in layer:
                    destination = get_destination_from_layer(layer)
                    if destination:
                        main_projet_polylines.append(polyline)
                        area = calculate_area(polyline)
                        
                        # Enregistrer les dimensions brutes pour analyse
                        vertices = polyline.get('vertices', [])
                        if vertices:
                            min_x = min(v['x'] for v in vertices)
                            max_x = max(v['x'] for v in vertices)
                            min_y = min(v['y'] for v in vertices)
                            max_y = max(v['y'] for v in vertices)
                            logger.info(f"Dimensions brutes pour {destination}: X({min_x},{max_x}), Y({min_y},{max_y}), Surface={area}")
                            
                        # Vérifier si le calcul semble correct (cas de figure spécifique)
                        # Les erreurs de calcul peuvent venir de points trop proches ou mal définis
                        
                        if destination not in calculation_results['projet']:
                            calculation_results['projet'][destination] = 0
                        calculation_results['projet'][destination] += area
                        logger.info(f"Projet: Destination {destination} - ajout surface {area}")
                elif is_special_layer(layer):
                    special_projet_polylines.append(polyline)
            
            # Déduire les surfaces spéciales des existantes
            for special_polyline in special_existant_polylines:
                area = calculate_area(special_polyline)
                if area <= 0:
                    continue
                    
                for main_polyline in main_existant_polylines:
                    if is_contained(special_polyline, main_polyline):
                        destination = get_destination_from_layer(main_polyline.get('layer', ''))
                        if destination and destination in calculation_results['existant']:
                            calculation_results['existant'][destination] -= area
                            logger.info(f"Existant: Déduction de {area} pour {destination}")
                        break
            
            # Déduire les surfaces spéciales des projets
            for special_polyline in special_projet_polylines:
                area = calculate_area(special_polyline)
                if area <= 0:
                    continue
                    
                for main_polyline in main_projet_polylines:
                    if is_contained(special_polyline, main_polyline):
                        destination = get_destination_from_layer(main_polyline.get('layer', ''))
                        if destination and destination in calculation_results['projet']:
                            calculation_results['projet'][destination] -= area
                            logger.info(f"Projet: Déduction de {area} pour {destination}")
                        break
            
            # IMPORTANT: Recalculer les surfaces projet avec la même méthode que pour les surfaces existantes
            # Nous allons vider et recalculer les surfaces projet
            logger.info("Recalcul des surfaces projet avec la même méthode que pour les surfaces existantes")
            
            # Vider les résultats précédents pour les surfaces projet
            calculation_results['projet'] = {}
            
            # Appliquer exactement la même méthode pour les surfaces projet que pour les surfaces existantes
            # Pour les polylignes principales
            main_projet_polylines = []
            for polyline in projet_polylines:
                layer = polyline.get('layer', '')
                if 'GEX_EDS_SDP_1' in layer:
                    destination = get_destination_from_layer(layer)
                    if destination:
                        main_projet_polylines.append(polyline)
                        area = calculate_area(polyline)
                        if destination not in calculation_results['projet']:
                            calculation_results['projet'][destination] = 0
                        calculation_results['projet'][destination] += area
                        logger.info(f"Recalcul Projet: Destination {destination} - ajout surface {area}")
            
            # Pour les polylignes spéciales
            special_projet_polylines = []
            for polyline in projet_polylines:
                layer = polyline.get('layer', '')
                if 'GEX_EDS_SDP_1' not in layer and is_special_layer(layer):
                    special_projet_polylines.append(polyline)
            
            # Déduire les surfaces spéciales des surfaces projet
            for special_polyline in special_projet_polylines:
                area = calculate_area(special_polyline)
                if area <= 0:
                    continue
                    
                # Vérifier la contenance dans chaque polyligne principale
                for main_polyline in main_projet_polylines:
                    if is_contained(special_polyline, main_polyline):
                        destination = get_destination_from_layer(main_polyline.get('layer', ''))
                        if destination and destination in calculation_results['projet']:
                            calculation_results['projet'][destination] -= area
                            logger.info(f"Recalcul Projet: Déduction de {area} pour {destination}")
                        break
            
            # Débogage des résultats de calcul finaux
            logger.info(f"===== Résultats de calcul finaux =====")
            for dest, value in calculation_results['existant'].items():
                logger.info(f"Existant - {dest}: {value}")
            for dest, value in calculation_results['projet'].items():
                logger.info(f"Projet - {dest}: {value}")
            
            # Collecter uniquement les destinations qui ont des valeurs dans les calculs
            all_destinations = set()
            
            # Nous n'ajoutons que les destinations qui ont des polylignes dans les fichiers extraits
            for destination in calculation_results['existant'].keys():
                all_destinations.add(destination)
                
            for destination in calculation_results['projet'].keys():
                all_destinations.add(destination)
                
            logger.info(f"Destinations présentes dans les fichiers extraits: {all_destinations}")
            
            # Vérifier si nous avons au moins une destination
            if not all_destinations:
                logger.warning("Aucune destination trouvée dans les fichiers extraits. Vérifiez le contenu des fichiers DXF.")
            
            # Traiter chaque destination
            for destination in sorted(all_destinations):
                formatted_destination = special_cases.get(destination, destination)
                
                # Surface existante (A) - du fichier Projet_demoli_feuille_TA.dxf
                existant_surface = calculation_results['existant'].get(destination, 0)
                
                # Surface projet - du fichier Existant_exmple_demoli.dxf - CONSERVER LA VALEUR BRUTE
                # Ne pas arrondir à ce stade pour avoir le même traitement que les surfaces existantes
                projet_surface = calculation_results['projet'].get(destination, 0)
                
                # Ajouter la ligne au fichier Excel avec précision complète pour les surfaces
                ws[f'B{row}'] = formatted_destination
                
                # Afficher les valeurs exactes dans les logs avant affichage
                logger.info(f"Destination {formatted_destination} - Surface existante: {existant_surface}, Surface projet: {projet_surface}")
                
                # Surface existante avec précision identique au fichier attendu
                if existant_surface > 0:
                    ws[f'C{row}'] = existant_surface
                
                # Surface projet avec précision identique au fichier attendu
                if projet_surface > 0:
                    ws[f'I{row}'] = projet_surface
                
                # Remplir la colonne Surface démolie (F) uniquement si > 0
                demolition_surface = calculation_results['demolition'].get(destination, 0)
                if demolition_surface > 0:
                    ws[f'F{row}'] = demolition_surface
                
                # Journaliser les résultats pour débogage
                logger.info(f"Destination {formatted_destination} - Surface existante: {existant_surface}, "
                           f"Surface projet: {projet_surface}, Surface démolie: {demolition_surface}")
                
                row += 1
            
            # Créer une nouvelle feuille nommée TA
            ws_ta = wb.create_sheet(title="TA")
            
            # En-têtes pour la feuille TA
            ws_ta['A1'] = "Etages"
            ws_ta['B1'] = "Destinations"
            ws_ta['C1'] = "TA existant"
            ws_ta['D1'] = "TA creee"
            ws_ta['E1'] = "TA demolie reconstruite"
            ws_ta['F1'] = "TA Supprimee"
            ws_ta['G1'] = "TA projet"
            ws_ta['H1'] = "TA pour Stationnement"
            
            # Remplir uniquement la colonne Etages dans la feuille TA (pas les autres colonnes)
            ta_row = 2
            ws_ta[f'A{ta_row}'] = floor_name  # Ajouter seulement l'étage
            
            # Laisser toutes les autres colonnes vides pour remplissage manuel
            # Juste une seule ligne avec l'étage
            
            # Sauvegarder le fichier Excel
            wb.save(excel_path)
            
            logger.info(f"Fichier Excel créé avec succès: {excel_path}")
        except Exception as e:
            logger.error(f"Erreur lors de la génération du fichier Excel: {str(e)}")
            return jsonify({'error': f'Erreur lors de la génération du fichier Excel: {str(e)}'}), 500
        
        # Réponse finale
        return jsonify({
            'message': 'Fichier Excel généré avec succès',
            'filePath': excel_path
        }), 201
    
    except Exception as e:
        # Gestionnaire général d'exceptions - attrape tout et journalise
        import traceback
        logger.error(f"ERREUR CRITIQUE dans generate_excel_file: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': f'Erreur lors de la génération du fichier Excel: {str(e)}'}), 500

@app.route('/download-excel-file', methods=['GET'])
def download_excel_file():
    """Permet le téléchargement d'un fichier Excel"""
    logger.info("Requête GET reçue pour télécharger un fichier Excel")
    
    try:
        file_path = request.args.get('filePath')
        email = request.args.get('email')
        
        if not file_path:
            logger.error("Chemin de fichier non fourni")
            return jsonify({'error': 'Chemin de fichier non fourni'}), 400
        
        if not email:
            logger.error("Email non fourni")
            return jsonify({'error': 'Email non fourni'}), 400
        
        logger.info(f"Fichier demandé: {file_path} pour l'utilisateur: {email}")
        
        # Extrait le nom du fichier du chemin (pour recherche directe)
        filename = os.path.basename(file_path)
        logger.info(f"Nom du fichier: {filename}")
        
        # Déterminer le dossier racine de l'utilisateur basé sur son email
        current_dir = os.path.dirname(os.path.abspath(__file__))
        logger.info(f"Répertoire courant: {current_dir}")
        
        folder_name = email.split('@')[0]
        
        # Essayons différentes combinaisons possibles pour le répertoire utilisateur
        possible_paths = [
            os.path.join(current_dir, 'app', 'Ressources', folder_name),
            os.path.join(current_dir, 'Ressources', folder_name),
        ]
        
        # Vérifier chaque chemin possible
        user_path_found = None
        for path in possible_paths:
            logger.info(f"Vérification du chemin: {path} (existe: {os.path.exists(path)})")
            if os.path.exists(path):
                user_path_found = path
                break
        
        if user_path_found is None:
            logger.error(f"Aucun dossier utilisateur trouvé pour {email}")
            return jsonify({'error': 'Dossier utilisateur non trouvé'}), 404
        
        # Premier essai: rechercher le fichier dans le chemin fourni
        full_file_path = os.path.join(user_path_found, file_path)
        logger.info(f"Tentative 1 - Chemin complet du fichier: {full_file_path}")
        
        # Si le fichier n'existe pas, rechercher récursivement dans le répertoire de l'utilisateur
        if not os.path.exists(full_file_path):
            logger.info("Recherche du fichier dans le répertoire utilisateur...")
            found_files = []
            
            # Recherche récursive du fichier
            for root, dirs, files in os.walk(user_path_found):
                if filename in files:
                    found_path = os.path.join(root, filename)
                    found_files.append(found_path)
                    logger.info(f"Fichier trouvé: {found_path}")
            
            if found_files:
                # Utiliser le premier fichier trouvé
                full_file_path = found_files[0]
                logger.info(f"Utilisation du fichier trouvé: {full_file_path}")
            else:
                logger.error(f"Aucun fichier {filename} trouvé dans le répertoire utilisateur")
                return jsonify({'error': 'Fichier non trouvé dans le répertoire utilisateur'}), 404
        
        if not os.path.exists(full_file_path):
            logger.error(f"Fichier non trouvé: {full_file_path}")
            return jsonify({'error': 'Fichier non trouvé'}), 404
            
        # Le fichier existe, vérifier sa taille
        file_size = os.path.getsize(full_file_path)
        logger.info(f"Taille du fichier: {file_size} octets")
        
        if file_size == 0:
            logger.error("Le fichier est vide")
            return jsonify({'error': 'Le fichier est vide'}), 500
        
        logger.info(f"Envoi du fichier: {full_file_path}")
        
        # Envoyer le fichier au client
        try:
            return send_file(
                full_file_path,
                as_attachment=True,
                download_name=os.path.basename(file_path),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            logger.error(f"Erreur lors de l'envoi du fichier: {str(e)}")
            return jsonify({'error': f"Erreur lors de l'envoi du fichier: {str(e)}"}), 500
    
    except Exception as e:
        logger.error(f"Erreur lors du téléchargement du fichier Excel: {str(e)}")
        return jsonify({'error': f'Erreur lors du téléchargement du fichier: {str(e)}'}), 500

def generate_visa_content(surfaces, floor_name):
    """Génère le contenu du fichier visa.txt avec tous les détails des calculs et des éléments"""
    now = datetime.datetime.now()
    content = [
        "=" * 80,
        f"RAPPORT DÉTAILLÉ DE CALCUL DE SURFACE - {floor_name}",
        "=" * 80,
        f"Date de génération: {now.strftime('%d/%m/%Y %H:%M:%S')}",
        f"Étage/Niveau: {floor_name}",
        "",
        "RÉSUMÉ DES SURFACES:",
        "-" * 50
    ]
    
    projet_surface = surfaces.get('projet', {}).get('surface', 0)
    content.append(f"Surface PROJET:\t\t{projet_surface:.2f} m²")
    
    existant_surface = surfaces.get('existant', {}).get('surface', 0)
    content.append(f"Surface EXISTANT:\t{existant_surface:.2f} m²")
    
    difference = surfaces.get('difference', 0)
    if existant_surface > 0:
        perc_diff = difference / existant_surface * 100
        content.append(f"DIFFÉRENCE:\t\t{difference:.2f} m² ({'+' if difference > 0 else ''}{perc_diff:.2f}% par rapport à l'existant)")
    else:
        content.append(f"DIFFÉRENCE:\t\t{difference:.2f} m²")
    content.append(f"IMPACT:\t\t\t{'Agrandissement' if difference > 0 else 'Réduction' if difference < 0 else 'Pas de changement'}")
    
    if projet_surface > 0 and surfaces.get('projet', {}).get('details', {}).get('polylines', 0) > 0:
        densite_projet = projet_surface / surfaces['projet']['details']['polylines']
        content.append(f"DENSITÉ PROJET:\t{densite_projet:.2f} m²/élément")
    if existant_surface > 0 and surfaces.get('existant', {}).get('details', {}).get('polylines', 0) > 0:
        densite_existant = existant_surface / surfaces['existant']['details']['polylines']
        content.append(f"DENSITÉ EXISTANT:\t{densite_existant:.2f} m²/élément")
    
    content.append("")
    content.append("ANALYSE COMPARATIVE DÉTAILLÉE:")  
    content.append("-" * 50)
    
    if projet_surface > 0 and existant_surface > 0:
        if difference > 0:
            content.append(f"Détails de l'agrandissement: {difference:.2f} m² ajoutés par rapport à l'existant")
            content.append(f"Ratio surface projet/existant: {projet_surface/existant_surface:.2f}")
            content.append(f"Pourcentage d'augmentation: {perc_diff:.2f}%")
            proj_elems = surfaces.get('projet', {}).get('details', {}).get('polylines', 0) + surfaces.get('projet', {}).get('details', {}).get('circles', 0)
            exist_elems = surfaces.get('existant', {}).get('details', {}).get('polylines', 0) + surfaces.get('existant', {}).get('details', {}).get('circles', 0)
            if proj_elems > exist_elems:
                content.append(f"Surface moyenne ajoutée par nouvel élément: {difference / (proj_elems - exist_elems):.2f} m²")
            
        elif difference < 0:
            content.append(f"Détails de la réduction: {abs(difference):.2f} m² supprimés par rapport à l'existant")
            content.append(f"Ratio surface projet/existant: {projet_surface/existant_surface:.2f}")
            content.append(f"Pourcentage de diminution: {abs(perc_diff):.2f}%")
            proj_elems = surfaces.get('projet', {}).get('details', {}).get('polylines', 0) + surfaces.get('projet', {}).get('details', {}).get('circles', 0)
            exist_elems = surfaces.get('existant', {}).get('details', {}).get('polylines', 0) + surfaces.get('existant', {}).get('details', {}).get('circles', 0)
            if exist_elems > proj_elems:
                content.append(f"Surface moyenne réduite par élément supprimé: {abs(difference) / (exist_elems - proj_elems):.2f} m²")
        else:
            content.append("Les surfaces projet et existant sont identiques")
            content.append("Aucune modification significative de la surface globale")
            proj_elems = surfaces.get('projet', {}).get('details', {}).get('polylines', 0) + surfaces.get('projet', {}).get('details', {}).get('circles', 0)
            exist_elems = surfaces.get('existant', {}).get('details', {}).get('polylines', 0) + surfaces.get('existant', {}).get('details', {}).get('circles', 0)
            if proj_elems != exist_elems:
                content.append(f"Bien que la surface totale soit identique, le nombre d'éléments a changé: {proj_elems - exist_elems:+d} éléments")
    elif projet_surface > 0 and existant_surface == 0:
        content.append(f"Nouvelle construction: {projet_surface:.2f} m² sans existant précédent")
        proj_elems = surfaces.get('projet', {}).get('details', {}).get('polylines', 0) + surfaces.get('projet', {}).get('details', {}).get('circles', 0)
        if proj_elems > 0:
            content.append(f"Surface moyenne par élément: {projet_surface / proj_elems:.2f} m²")
    elif projet_surface == 0 and existant_surface > 0:
        content.append(f"Démolition complète: {existant_surface:.2f} m² de surface existante")
        exist_elems = surfaces.get('existant', {}).get('details', {}).get('polylines', 0) + surfaces.get('existant', {}).get('details', {}).get('circles', 0)
        if exist_elems > 0:
            content.append(f"Surface moyenne par élément supprimé: {existant_surface / exist_elems:.2f} m²")
        
    content.append("")
    content.append("INVENTAIRE DÉTAILLÉ DES ÉLÉMENTS:")
    content.append("-" * 50)
    
    if surfaces.get('projet', {}).get('details'):
        projet_details = surfaces['projet']['details']
        content.append("PROJET:")
        content.append(f"  - Nombre total de polylignes: {projet_details.get('polylines', 0)}")
        content.append(f"  - Nombre total de cercles: {projet_details.get('circles', 0)}")
        content.append(f"  - Nombre total d'éléments: {projet_details.get('polylines', 0) + projet_details.get('circles', 0)}")
        if 'polylines' in projet_details and projet_details['polylines'] > 0:
            content.append(f"  - Surface moyenne par polyligne: {projet_surface / projet_details['polylines']:.2f} m² (si répartition uniforme)")
        
        content.append("  - Types d'éléments: Polylignes fermées (surfaces), cercles")
        if projet_details.get('polylines', 0) > 0 and projet_details.get('circles', 0) > 0:
            content.append(f"  - Ratio polylignes/cercles: {projet_details.get('polylines', 0) / projet_details.get('circles', 1):.2f}")
        
    if surfaces.get('existant', {}).get('details'):
        existant_details = surfaces['existant']['details']
        content.append("\nEXISTANT:")
        content.append(f"  - Nombre total de polylignes: {existant_details.get('polylines', 0)}")
        content.append(f"  - Nombre total de cercles: {existant_details.get('circles', 0)}")
        content.append(f"  - Nombre total d'éléments: {existant_details.get('polylines', 0) + existant_details.get('circles', 0)}")
        if 'polylines' in existant_details and existant_details['polylines'] > 0:
            content.append(f"  - Surface moyenne par polyligne: {existant_surface / existant_details['polylines']:.2f} m² (si répartition uniforme)")
        
        content.append("  - Types d'éléments: Polylignes fermées (surfaces), cercles")
        if existant_details.get('polylines', 0) > 0 and existant_details.get('circles', 0) > 0:
            content.append(f"  - Ratio polylignes/cercles: {existant_details.get('polylines', 0) / existant_details.get('circles', 1):.2f}")
    
    if surfaces.get('projet', {}).get('details') and surfaces.get('existant', {}).get('details'):
        projet_details = surfaces['projet']['details']
        existant_details = surfaces['existant']['details']
        
        content.append("\nCOMPARAISON DES ÉLÉMENTS:")
        poly_diff = projet_details.get('polylines', 0) - existant_details.get('polylines', 0)
        circle_diff = projet_details.get('circles', 0) - existant_details.get('circles', 0)
        total_elem_diff = poly_diff + circle_diff
        
        content.append(f"  - Différence en polylignes: {'+' if poly_diff > 0 else ''}{poly_diff}")
        content.append(f"  - Différence en cercles: {'+' if circle_diff > 0 else ''}{circle_diff}")
        content.append(f"  - Différence totale d'éléments: {'+' if total_elem_diff > 0 else ''}{total_elem_diff}")
    
    content.append("")
    content.append("COMMENTAIRES ET OBSERVATIONS:")
    content.append("-" * 50)
    
    if difference > 0:
        content.append(f"- Ce calcul montre une augmentation de surface de {difference:.2f} m² ({perc_diff:.2f}% par rapport à l'existant).")
        content.append("- Veuillez vérifier que cette augmentation est conforme aux règles d'urbanisme et aux limites de constructibilité.")
        content.append(f"- Vérification recommandée: coefficient d'emprise au sol (CES) et coefficient d'occupation des sols (COS) du PLU.")
        
        content.append("- Points d'attention:")
        content.append("  * Vérifier les calculs des droits à construire dans le cas d'une extension")
        content.append("  * Contrôler la conformité avec les règles de constructibilité locales")
        content.append("  * Confirmer la compatibilité avec les règles de gabarit et de prospect")
        
        if perc_diff > 20:
            content.append(f"- ATTENTION: L'augmentation de surface de {perc_diff:.2f}% est significative et peut nécessiter ")
            content.append("  des autorisations d'urbanisme spécifiques (permis de construire plutôt qu'une déclaration préalable).")
        
    elif difference < 0:
        content.append(f"- Ce calcul montre une diminution de surface de {abs(difference):.2f} m² ({abs(perc_diff):.2f}% par rapport à l'existant).")
        content.append("- Assurez-vous que cette réduction est voulue et qu'elle respecte les objectifs du projet.")
        
        content.append("- Points d'attention:")
        content.append("  * Vérifier l'impact sur le fonctionnement des espaces")
        content.append("  * Contrôler la conformité des nouveaux espaces avec les normes d'accessibilité")
        content.append("  * Évaluer l'impact sur les calculs réglementaires (surface utile, surface taxable)")
        
        if abs(perc_diff) > 30:
            content.append(f"- NOTE: La réduction importante de {abs(perc_diff):.2f}% peut indiquer une restructuration majeure")
            content.append("  du bâtiment. Vérifiez l'exactitude des fichiers DXF et des éléments pris en compte.")
    else:
        content.append("- Les surfaces projet et existant sont identiques.")
        content.append("- Aucun impact sur la surface totale.")
        
        proj_elems = surfaces.get('projet', {}).get('details', {}).get('polylines', 0) + surfaces.get('projet', {}).get('details', {}).get('circles', 0)
        exist_elems = surfaces.get('existant', {}).get('details', {}).get('polylines', 0) + surfaces.get('existant', {}).get('details', {}).get('circles', 0)
        if proj_elems != exist_elems:
            content.append(f"- Bien que la surface totale soit inchangée, la répartition des espaces a été modifiée")
            content.append("  (nombre d'éléments différent entre projet et existant).")
            content.append("- Vérifier que cette redistribution respecte les exigences fonctionnelles du bâtiment.")
        else:
            content.append("- La structure du bâtiment semble inchangée (même nombre d'éléments).")
            content.append("- Il peut s'agir d'une mise à jour du fichier sans modification structurelle.")
    
    content.append("")
    content.append("RECOMMANDATIONS TECHNIQUES:")
    content.append("-" * 50)
    content.append("- Vérification de la cohérence des surfaces avec les autres documents du projet")
    content.append("- Confirmation des calculs avec les surfaces mentionnées dans les dossiers administratifs")
    content.append("- Évaluation de l'impact des modifications sur les performances énergétiques du bâtiment")
    content.append("- Analyse des implications sur les évacuations et issues de secours (si applicable)")
    
    content.append("")
    content.append("=" * 80)
    content.append("INFORMATIONS LÉGALES:")
    content.append("-" * 50)
    content.append(f"Ce rapport a été généré automatiquement le {now.strftime('%d/%m/%Y')} à {now.strftime('%H:%M:%S')}")
    content.append("Les calculs sont effectués sur la base des éléments fournis dans les fichiers DXF.")
    content.append("Ces résultats doivent être vérifiés par un professionnel qualifié.")
    content.append("")
    return "\n".join(content)

def create_excel_document(excel_path, surfaces, floor_name):
    """Crée un document Excel avec les données de surface"""
    wb = openpyxl.Workbook()
    
    # Configuration des styles communs
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    data_font = Font(name='Arial', size=12)
    
    title_alignment = Alignment(horizontal='center', vertical='center')
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    # Création de la feuille TA
    ws_ta = wb.active
    ws_ta.title = "TA"
    
    # En-têtes colonnes TA
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    for col in columns:
        ws_ta.column_dimensions[col].width = 15
    
    # En-tête de la feuille TA - modifié selon les demandes
    ws_ta['A1'] = "Étages"
    ws_ta['B1'] = "TA existant"
    ws_ta['C1'] = "TA creee"
    ws_ta['D1'] = "TA demolie reconstruite"
    ws_ta['E1'] = "TA Supprimee"
    ws_ta['F1'] = "TA projet"
    ws_ta['G1'] = "TA pour Stationnement"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = ws_ta[f'{col}1']
        cell.font = subheader_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Ajouter uniquement le nom d'étage dans la feuille TA
    ws_ta['A2'] = floor_name
    
    # Appliquer le style uniquement à la cellule du nom d'étage
    cell = ws_ta['A2']
    cell.font = data_font
    cell.alignment = data_alignment
    cell.border = border
    
    # Création de la feuille SDP
    ws_sdp = wb.create_sheet(title="SDP")
    
    # Configuration des colonnes SDP
    for col in columns:
        ws_sdp.column_dimensions[col].width = 15
    
    # En-tête de la feuille SDP - avec les titres modifiés et les colonnes divisées
    ws_sdp['A1'] = "Etages"
    ws_sdp['B1'] = "Destinations"
    ws_sdp['C1'] = "Surface existante avant travaux (A)"
    ws_sdp['D1'] = "Surface creee (B)"
    ws_sdp['E1'] = "Surface creee par changement de destination (C)"
    ws_sdp['F1'] = "Surface demolie reconstruite"
    ws_sdp['G1'] = "Surface supprimee (D)"
    ws_sdp['H1'] = "Surface supprimee par changement de destination"
    ws_sdp['I1'] = "Surface projet"
    ws_sdp['J1'] = "Surface RDV"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        cell = ws_sdp[f'{col}1']
        cell.font = subheader_font
        cell.alignment = header_alignment
        cell.border = border
        
    # Ajouter le nom d'étage dans la feuille SDP
    ws_sdp['A2'] = floor_name
    
    # Appliquer le style à la cellule du nom d'étage
    cell = ws_sdp['A2']
    cell.font = data_font
    cell.alignment = data_alignment
    cell.border = border
    
    # Utiliser les données réelles des polylignes pour la colonne Destinations
    row = 2  # Commencer à la ligne 2
    
    # Fonction pour extraire le nom de destination à partir d'une polyligne
    def get_destination_name(polyline):
        import re
        
        # Essayer d'extraire le nom du calque (layer)
        layer = polyline.get('layer', '')
        
        # Si le calque n'est pas valide, essayer d'autres propriétés
        if not layer or not isinstance(layer, str):
            # Essayer d'utiliser le type de polyligne
            polyline_type = polyline.get('type', '')
            if polyline_type:
                return polyline_type
            return "NON_SPECIFIE"
        
        # Format standard: GEX_EDS_SDP_X-CATEGORIE_SOUSCATEGORIE
        # Exemple: GEX_EDS_SDP_1-COMMERCE_CIN -> COMMERCE_CIN
        pattern1 = r'(?:GEX_EDS_)?SDP_\d+-([A-Z_]+)'
        match = re.search(pattern1, layer)
        if match:
            # Extraire uniquement la partie après le tiret (COMMERCE_CIN)
            return match.group(1)
        
        # Format spécifique pour hauteur sous plafond (H-180)
        if "H-" in layer:
            match = re.search(r'H-(\d+)', layer)
            if match:
                return f"H-{match.group(1)}"
        
        # Format spécifique pour stationnement
        if "PK" in layer:
            return "PK"
        
        # Si le calque contient un tiret, prendre la partie après le tiret
        if "-" in layer:
            parts = layer.split('-')
            if len(parts) > 1:
                return parts[-1]
        
        # Enlever les préfixes communs
        clean_layer = re.sub(r'^(?:GEX_EDS_)?(?:SDP_)?\d*', '', layer)
        
        # Si le calque nettoyé commence par un tiret, l'enlever
        if clean_layer.startswith('-'):
            clean_layer = clean_layer[1:]
        
        # Si le calque nettoyé n'est pas vide, le retourner
        if clean_layer:
            return clean_layer
        
        # En dernier recours, retourner le calque original
        return layer
    
    # Collecter les noms de destination uniques et calculer les surfaces par destination
    destinations = set()  # Utiliser un ensemble pour éliminer automatiquement les doublons
    
    # Créer un dictionnaire pour stocker tous les types de calques rencontrés
    all_layer_types = {}
    
    # Fonction pour enregistrer un type de calque
    def register_layer_type(polyline):
        layer = polyline.get('layer', '')
        if layer and isinstance(layer, str):
            # Extraire le nom de destination
            destination = get_destination_name(polyline)
            all_layer_types[destination] = True
    
    # Dictionnaire pour stocker les surfaces par destination
    surfaces_par_destination = {}
    
    # Extraire les destinations des polylignes du projet
    projet_polylines = surfaces.get('projet', {}).get('polylines', [])
    for polyline in projet_polylines:
        destination = get_destination_name(polyline)
        destinations.add(destination)
    
    # Fonction pour calculer la surface d'une polyligne à partir de ses vertices
    def calculer_surface_polyligne(polyline):
        # Vérifier si la polyligne a des vertices
        if 'vertices' not in polyline or not polyline['vertices']:
            return 0
        
        vertices = polyline['vertices']
        if len(vertices) < 3:  # Un polygone doit avoir au moins 3 points
            return 0
        
        # Calculer la surface en utilisant la formule de Shoelace (Gauss's area formula)
        area = 0.0
        n = len(vertices)
        for i in range(n):
            j = (i + 1) % n
            area += vertices[i]['x'] * vertices[j]['y']
            area -= vertices[j]['x'] * vertices[i]['y']
        
        # Prendre la valeur absolue et diviser par 2
        return abs(area) / 2.0
    
    # IMPORTANT: Selon la demande du client et en tenant compte de l'inversion des fichiers
    # La colonne "Surface existante avant travaux (A)" contient les surfaces des fichiers de "Projet_demoli_feuille_TA.dxf" (section "Importation d'existant .dxf")
    # La colonne "Surface projet" contient les surfaces des fichiers de "Existant_exmple_demoli.dxf" (section "Importation de fichier .dxf")
    
    # Fonction pour déterminer si une polyligne est un calque principal SDP_1
    def is_main_sdp_polyline(polyline):
        layer = polyline.get('layer', '')
        return isinstance(layer, str) and layer.startswith('GEX_EDS_SDP_1')
    
    # Fonction pour déterminer si une polyligne est un calque spécial à déduire (SDP_2, SDP_3, etc.)
    def is_special_polyline(polyline):
        layer = polyline.get('layer', '')
        if not isinstance(layer, str):
            return False
        # Vérifier si c'est un calque spécial à déduire (pas SDP_1)
        patterns = ['GEX_EDS_SDP_2', 'GEX_EDS_SDP_3', 'GEX_EDS_SDP_4', 'GEX_EDS_SDP_5', 'GEX_EDS_SDP_7']
        return any(pattern in layer for pattern in patterns)
    
    # Fonction pour déterminer la destination parente d'un calque spécial de manière simple sans dépendre de Shapely
    def get_parent_destination(special_polyline, main_polylines):
        special_layer = special_polyline.get('layer', '')
        if not isinstance(special_layer, str):
            return None
            
        # Méthode simplifiée: utiliser les correspondances de noms de calques
        # Par exemple, si nous avons GEX_EDS_SDP_3-H-180, chercher un GEX_EDS_SDP_1 avec des coordonnées similaires
        special_prefix = special_layer.split('-')[0] if '-' in special_layer else ''
        
        # Essayer d'extraire le numéro d'étage ou de bâtiment (si présent dans le nom du calque)
        parts = special_prefix.split('_')
        etage_num = parts[-1] if len(parts) > 1 and parts[-1].isdigit() else None
        
        # Vérifier chaque polyligne principale pour voir si elle pourrait être parent
        # basé sur des heuristiques simples (proximité ou même étage)
        for main_polyline in main_polylines:
            main_layer = main_polyline.get('layer', '')
            if not isinstance(main_layer, str):
                continue
                
            # Si le calque spécial est inclus dans les coordonnées du calque principal, c'est probablement son parent
            if is_contained(special_polyline, main_polyline):
                return get_destination_name(main_polyline)
                
        # Si nous n'avons pas pu trouver de parent, retourner None
        return None
        
    # Fonction helper pour déterminer si une polyligne est contenue dans une autre
    # en utilisant une méthode simplifiée de comparaison de boîtes englobantes
    def is_contained(polyline1, polyline2):
        try:
            # Essayer de déterminer si polyline1 est contenue dans polyline2
            # en vérifiant si les coordonnées de polyline1 sont à l'intérieur de la boîte englobante de polyline2
            vertices1 = polyline1.get('vertices', [])
            vertices2 = polyline2.get('vertices', [])
            
            if not vertices1 or not vertices2:
                return False
                
            # Trouver les limites de la boîte englobante pour polyline2
            min_x2 = min(v['x'] for v in vertices2)
            max_x2 = max(v['x'] for v in vertices2)
            min_y2 = min(v['y'] for v in vertices2)
            max_y2 = max(v['y'] for v in vertices2)
            
            # Vérifier si tous les points de polyline1 sont à l'intérieur de la boîte englobante de polyline2
            points_inside = 0
            for v in vertices1:
                if min_x2 <= v['x'] <= max_x2 and min_y2 <= v['y'] <= max_y2:
                    points_inside += 1
            
            # Si au moins la moitié des points sont à l'intérieur, considérer comme contenu
            return points_inside >= len(vertices1) / 2
        except Exception as e:
            logger.warning(f"Erreur lors de la vérification de contenance: {str(e)}")
            return False
    
    # Récupérer les polylignes des deux fichiers
    existant_polylines = surfaces.get('existant', {}).get('polylines', [])
    projet_polylines = surfaces.get('projet', {}).get('polylines', [])
    
    # Initialiser le dictionnaire des surfaces
    surfaces_par_destination = {}
    
    # Filtrer les polylignes principales (SDP_1)
    main_existant_polylines = [p for p in existant_polylines if is_main_sdp_polyline(p)]
    main_projet_polylines = [p for p in projet_polylines if is_main_sdp_polyline(p)]
    
    # Filtrer les polylignes spéciales à déduire
    special_existant_polylines = [p for p in existant_polylines if is_special_polyline(p)]
    special_projet_polylines = [p for p in projet_polylines if is_special_polyline(p)]
    
    # Collecter toutes les destinations des polylignes principales
    all_destinations = set()
    for polyline in main_existant_polylines + main_projet_polylines:
        destination = get_destination_name(polyline)
        all_destinations.add(destination)
    
    # Initialiser le dictionnaire des surfaces pour toutes les destinations
    for destination in all_destinations:
        surfaces_par_destination[destination] = {'existant': 0, 'projet': 0, 'rdv': 0}
    
    # CORRECTION: Calcul des surfaces existantes (colonne "Surface existante avant travaux (A)")
    # Utiliser les polylignes de "Projet_demoli_feuille_TA.dxf" (dans main_existant_polylines)
    for polyline in main_existant_polylines:
        destination = get_destination_name(polyline)
        surface = calculer_surface_polyligne(polyline)
        surfaces_par_destination[destination]['existant'] += surface
    
    # CORRECTION: Calcul des surfaces projet (colonne "Surface projet")
    # Utiliser les polylignes de "Existant_exmple_demoli.dxf" (dans main_projet_polylines)
    for polyline in main_projet_polylines:
        destination = get_destination_name(polyline)
        surface = calculer_surface_polyligne(polyline)
        surfaces_par_destination[destination]['projet'] += surface
    
    # Méthode simplifiée pour déduire les surfaces des calques spéciaux
    logger.info("Début de la déduction des surfaces des calques spéciaux")
    
    # Pour les fichiers DXF, on sait que GEX_EDS_SDP_2, GEX_EDS_SDP_3, etc. sont des zones à déduire
    # des zones principales GEX_EDS_SDP_1 correspondantes
    
    # Structure pour garder trace des calques spéciaux traités
    special_layers_processed = {
        'existant': {},
        'projet': {}
    }
    
    # Traitement des polylignes spéciales dans le fichier existant
    try:
        logger.info(f"Traitement de {len(special_existant_polylines)} polylignes spéciales dans le fichier existant")
        
        for special_polyline in special_existant_polylines:
            special_layer = special_polyline.get('layer', '')
            if not isinstance(special_layer, str):
                continue
                
            logger.info(f"Traitement du calque spécial existant: {special_layer}")
            
            # Calculer la surface de cette polyligne spéciale
            special_surface = calculer_surface_polyligne(special_polyline)
            if special_surface <= 0:
                logger.warning(f"Surface nulle ou négative pour le calque {special_layer}, ignoré")
                continue
                
            # Déterminer le type de calque spécial (SDP_2, SDP_3, etc.)
            if 'GEX_EDS_SDP_2' in special_layer:
                category = 'TREMIE'
            elif 'GEX_EDS_SDP_3' in special_layer:
                category = 'H-180'
            elif 'GEX_EDS_SDP_5' in special_layer:
                category = 'PK'
            elif 'GEX_EDS_SDP_7' in special_layer:
                category = 'LT'
            else:
                logger.warning(f"Type de calque spécial inconnu: {special_layer}, ignoré")
                continue
                
            # Essayer de trouver la destination parente directement via les coordonnées
            # Cette méthode est plus simple et plus robuste
            parent_found = False
            
            for main_polyline in main_existant_polylines:
                if is_contained(special_polyline, main_polyline):
                    parent_destination = get_destination_name(main_polyline)
                    if parent_destination in surfaces_par_destination:
                        # Déduire la surface spéciale de la surface existante
                        surfaces_par_destination[parent_destination]['existant'] -= special_surface
                        logger.info(f"Surface {special_surface} déduite de {parent_destination} (existant) - Calque: {special_layer}")
                        special_layers_processed['existant'][special_layer] = {
                            'surface': special_surface,
                            'parent': parent_destination
                        }
                        parent_found = True
                        break
            
            if not parent_found:
                logger.warning(f"Aucun parent trouvé pour le calque {special_layer} dans l'existant")
                
    except Exception as e:
        logger.error(f"Erreur lors du traitement des calques spéciaux existants: {str(e)}")
    
    # Traitement des polylignes spéciales dans le fichier projet
    try:
        logger.info(f"Traitement de {len(special_projet_polylines)} polylignes spéciales dans le fichier projet")
        
        for special_polyline in special_projet_polylines:
            special_layer = special_polyline.get('layer', '')
            if not isinstance(special_layer, str):
                continue
                
            logger.info(f"Traitement du calque spécial projet: {special_layer}")
            
            # Calculer la surface de cette polyligne spéciale
            special_surface = calculer_surface_polyligne(special_polyline)
            if special_surface <= 0:
                logger.warning(f"Surface nulle ou négative pour le calque {special_layer}, ignoré")
                continue
                
            # Déterminer le type de calque spécial (SDP_2, SDP_3, etc.)
            if 'GEX_EDS_SDP_2' in special_layer:
                category = 'TREMIE'
            elif 'GEX_EDS_SDP_3' in special_layer:
                category = 'H-180'
            elif 'GEX_EDS_SDP_5' in special_layer:
                category = 'PK'
            elif 'GEX_EDS_SDP_7' in special_layer:
                category = 'LT'
            else:
                logger.warning(f"Type de calque spécial inconnu: {special_layer}, ignoré")
                continue
                
            # Essayer de trouver la destination parente directement via les coordonnées
            parent_found = False
            
            for main_polyline in main_projet_polylines:
                if is_contained(special_polyline, main_polyline):
                    parent_destination = get_destination_name(main_polyline)
                    if parent_destination in surfaces_par_destination:
                        # Déduire la surface spéciale de la surface projet
                        surfaces_par_destination[parent_destination]['projet'] -= special_surface
                        logger.info(f"Surface {special_surface} déduite de {parent_destination} (projet) - Calque: {special_layer}")
                        special_layers_processed['projet'][special_layer] = {
                            'surface': special_surface,
                            'parent': parent_destination
                        }
                        parent_found = True
                        break
            
            if not parent_found:
                logger.warning(f"Aucun parent trouvé pour le calque {special_layer} dans le projet")
                
    except Exception as e:
        logger.error(f"Erreur lors du traitement des calques spéciaux projet: {str(e)}")
    
    # Afficher un résumé des calques spéciaux traités
    logger.info(f"Résumé des calques spéciaux traités - Existant: {len(special_layers_processed['existant'])}, Projet: {len(special_layers_processed['projet'])}")

        
    # Importer les bibliothèques nécessaires pour les analyses spatiales
    try:
        from shapely.geometry import Polygon, Point
        shapely_available = True
    except ImportError:
        # Si Shapely n'est pas disponible, on affiche un message
        print("La bibliothèque Shapely n'est pas disponible. L'analyse spatiale précise ne sera pas possible.")
        shapely_available = False
        
    # Fonction pour convertir une polyligne en polygone Shapely
    def polyligne_to_polygon(polyline):
        if 'vertices' not in polyline or not polyline['vertices'] or len(polyline['vertices']) < 3:
            return None
        vertices = polyline['vertices']
        # Créer une liste de coordonnées (x, y) pour Shapely
        coords = [(vertex['x'], vertex['y']) for vertex in vertices]
        try:
            # Créer un polygone Shapely
            return Polygon(coords)
        except Exception as e:
            print(f"Erreur lors de la création du polygone: {e}")
            return None
    
    # Identifier les polylignes LOC_SOC et SANITAIRES dans tous les fichiers
    rdv_specific_polylines = []
    for polyline in projet_polylines + existant_polylines:
        layer = polyline.get('layer', '')
        if isinstance(layer, str) and ('LOC_SOC' in layer or 'SANITAIRES' in layer):
            rdv_specific_polylines.append(polyline)
    
    # Pour chaque destination (polyligne SDP), vérifier si elle contient des polylignes LOC_SOC ou SANITAIRES
    if shapely_available and rdv_specific_polylines:
        for destination in all_destinations:
            # Trouver toutes les polylignes associées à cette destination
            destination_polylines = []
            for polyline in filtered_projet_polylines + filtered_existant_polylines:
                if get_destination_name(polyline) == destination:
                    destination_polylines.append(polyline)
            
            # Pour chaque polyligne de destination, vérifier si elle contient des polylignes LOC_SOC ou SANITAIRES
            for dest_polyline in destination_polylines:
                dest_polygon = polyligne_to_polygon(dest_polyline)
                if not dest_polygon:
                    continue
                
                # Vérifier chaque polyligne spécifique
                for rdv_polyline in rdv_specific_polylines:
                    rdv_polygon = polyligne_to_polygon(rdv_polyline)
                    if not rdv_polygon:
                        continue
                    
                    # Vérifier si le polygone RDV est contenu dans le polygone de destination
                    if dest_polygon.contains(rdv_polygon) or dest_polygon.intersects(rdv_polygon):
                        # Calcul de la surface du polygone RDV
                        surface_rdv = rdv_polygon.area
                        # Ajouter cette surface à la destination
                        surfaces_par_destination[destination]['rdv'] += surface_rdv
        
    # Utiliser toutes les destinations collectées pour l'affichage
    destinations = all_destinations
    
    # Fonction pour formater les noms de destination pour un affichage plus lisible
    def format_destination(destination):
        # Dictionnaire de correspondance pour les cas spéciaux
        special_cases = {
            "AUTRE_BUREAU": "Autre bureau",
            "AUTRE_CONGRE": "Autre congrès exposition",
            "AUTRE_ENTREP": "Autre entrepôt",
            "AUTRE_INDUST": "Autre industrie",
            "COMMERCE_ART": "Commerce artisanat",
            "COMMERCE_AUT": "Commerce autre hébergement touristique",
            "COMMERCE_CIN": "Commerce cinéma",
            "COMMERCE_DE_": "Commerce de gros",
            "COMMERCE_HOT": "Commerce hôtel",
            "COMMERCE_RES": "Commerce restauration",
            "COMMERCE_SER": "Commerce service accueil clientèle",
            "EXPLOITATIO": "Exploitation forestière",
            "EXPLOITATION": "Exploitation agricole",
            "HABITATION_H": "Habitation hébergement",
            "HABITATION_L": "Habitation logement",
            "SPIC_ADMINIS": "Spic administration",
            "SPIC_ART_SPE": "Spic art spectacle",
            "SPIC_AUTRE": "Spic autre",
            "SPIC_ENSEIGN": "Spic enseignement santé",
            "SPIC_LT": "Spic lt",
            "SPIC_SPORT": "Spic sport"
        }
        
        # Si la destination est dans notre dictionnaire, utiliser la correspondance
        if destination in special_cases:
            return special_cases[destination]
        
        # Sinon, transformer le format générique: remplacer les underscores par des espaces et mettre la première lettre en majuscule
        formatted = destination.replace('_', ' ').lower().capitalize()
        return formatted
    
    # Trier les destinations par ordre alphabétique
    sorted_destinations = sorted(destinations)
    
    # Ajouter les destinations uniques à la feuille SDP avec leurs surfaces
    row = 2  # Commencer à la ligne 2
    for destination in sorted_destinations:
        # Colonne B: Destination (formatée pour une meilleure lisibilité)
        formatted_destination = format_destination(destination)
        ws_sdp[f'B{row}'] = formatted_destination
        
        # Appliquer le style à la cellule
        cell = ws_sdp[f'B{row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Colonne C: Surface existante avant travaux (A)
        # Cette colonne est maintenant laissée intentionnellement vide selon la demande
        cell = ws_sdp[f'C{row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Colonne I: Surface projet
        # Cette colonne est maintenant laissée intentionnellement vide selon la demande
        cell = ws_sdp[f'I{row}']
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border
        
        # Colonne J: Surface RDV
        surface_rdv = surfaces_par_destination.get(destination, {}).get('rdv', 0)
        if surface_rdv > 0:
            # Convertir en mètres carrés si nécessaire
            surface_rdv_m2 = surface_rdv
            
            # Arrondir à 2 décimales
            ws_sdp[f'J{row}'] = round(surface_rdv_m2, 2)
            
            # Appliquer le style à la cellule
            cell = ws_sdp[f'J{row}']
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # Ajouter un commentaire indiquant la source de la donnée
            cell.comment = Comment(f"Surface calculée à partir des polylignes RDV (GEX_EDS_RDV_)", "Calcul automatique")
            cell.comment.width = 300
            cell.comment.height = 50
        
        row += 1
    
    # Sauvegarde du fichier Excel
    wb.save(excel_path)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)